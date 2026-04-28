#!/usr/bin/env python3
"""
사방넷 풀필먼트 발주등록 엑셀 생성 - 최종 통합본 v2
====================================================

입력 포맷 (자동 감지):
  1. 사방넷 관리자 주문 JSON (ordNo, clctPrdNm, clctSkuNm, ecptRmteNm ...)
  2. 스마트스토어 주문 엑셀 (.xlsx: 상품번호, 옵션정보, 수취인명 ...)

매핑 우선순위:
  1. 상품번호별_매핑 (스마트스토어 입력에서만 사용, 색상+사이즈 완전 일치)
  2. 키워드_매핑 (상품명/옵션을 배열 순서대로 훑어 최초 매칭 사용)
  3. 실패시 unmapped 리포트

세트 수량 변환:
  옵션 문자열에 "1+1"/"2+2"/"3+3" 포함 → 세트_배수 곱함
  옵션 문자열에 "N개" 포함 → N 곱함

사업자 분리 출력:
  E* 코드 = 이더컴퍼니 (공산품)  → <base>_ether.xlsx
  N* 코드 = 뉴트리정 (영양제)    → <base>_nutri.xlsx
  미분류                          → <base>_unmapped.xlsx

Usage:
  python3 create_fulfillment_excel.py \\
    --orders orders.json \\
    --mapping product_mapping.json \\
    --output fulfillment_YYYYMMDD.xlsx
"""

import argparse
import json
import os
import re
import sys

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
except ImportError:
    print("openpyxl 설치 필요: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


HEADERS = [
    "상품고유코드", "판매상품명", "수량", "배송방식",
    "주문자 이름", "받는분 이름", "전화번호1", "전화번호2",
    "우편번호", "주소1", "주소2", "배송메세지",
    "주문번호", "관리메모1", "관리메모2", "관리메모3",
    "관리메모4", "관리메모5", "상품별 메모1", "상품별 메모2",
    "상품별 메모3", "발주 타입", "출고희망일",
]
REQUIRED_HEADERS = {
    "상품고유코드", "판매상품명", "수량", "배송방식",
    "받는분 이름", "전화번호1", "우편번호", "주소1",
}

SABANG_FIELD_MAP = {
    "ordNo": "ordNo",
    "shmaOrdNo": "shmaOrdNo",
    "clctPrdNm": "product",
    "clctSkuNm": "option",
    "ordQty": "qty",
    "ecptRmteNm": "receiver",
    "ecptRmteTelNo": "phone",
    "rmteZipcd": "zipcode",
    "ecptRmteTotAddr": "address",
    "shpmtEtcFldVl": "message",
}


def load_mapping(path):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def match_by_product_no(mapping, product_no, color, size):
    pm = mapping.get("상품번호별_매핑", {})
    info = pm.get(str(product_no))
    if not info:
        return None
    if info.get("색상필요"):
        if not color or not size:
            return None
        key_raw = f"{color}_{size}"
        if key_raw in info.get("매핑", {}):
            key = key_raw
        else:
            color_conv = info.get("색상변환", {})
            normalized = color_conv.get(color, color)
            key = f"{normalized}_{size}"
    else:
        if not size:
            return None
        key = size
    code = info.get("매핑", {}).get(key)
    if not code:
        return None
    name = mapping.get("코드별_상품명", {}).get(code, "")
    return code, name, info.get("skip_multiplier", False)


def _all(text, kws):
    return all(k.lower() in text for k in (kws or []))


def _any(text, kws):
    return any(k.lower() in text for k in (kws or []))


def match_by_keyword(mapping, product, option):
    text = f"{product} {option}".lower()
    size_m = re.search(r"(2[2-8]\d)\s*[-~]\s*(2[2-8]\d)", text)
    size_key = f"{size_m.group(1)}-{size_m.group(2)}" if size_m else None

    for rule in mapping.get("키워드_매핑", []):
        # GUARD: keywords_all/keywords_any 둘 다 비면 룰 스킵 (silent failure 방지)
        ka = rule.get("keywords_all") or []
        ky = rule.get("keywords_any") or []
        if not ka and not ky:
            continue
        if ka and not _all(text, ka):
            continue
        if ky and not _any(text, ky):
            continue

        skip_mult = rule.get("skip_multiplier", False)
        if "옵션_분기" in rule:
            opt_low = option.lower()
            for br in rule["옵션_분기"]:
                if "옵션_포함_모두" in br and all(k.lower() in opt_low for k in br["옵션_포함_모두"]):
                    return br["code"], mapping["코드별_상품명"].get(br["code"], ""), br.get("skip_multiplier", skip_mult)
                if "옵션_포함" in br and br["옵션_포함"].lower() in opt_low:
                    return br["code"], mapping["코드별_상품명"].get(br["code"], ""), br.get("skip_multiplier", skip_mult)
            continue

        if "색상_사이즈_매핑" in rule:
            color_conv = rule.get("색상변환", {})
            color_norm = None
            for raw, norm in color_conv.items():
                if raw in text:
                    color_norm = norm
                    break
            if not color_norm or not size_key:
                continue
            code = rule["색상_사이즈_매핑"].get(f"{color_norm}_{size_key}")
            if code:
                return code, mapping["코드별_상품명"].get(code, ""), skip_mult
            continue

        if "사이즈_매핑" in rule:
            if not size_key:
                continue
            code = rule["사이즈_매핑"].get(size_key)
            if code:
                return code, mapping["코드별_상품명"].get(code, ""), skip_mult
            continue

        if "code" in rule:
            code = rule["code"]
            return code, mapping["코드별_상품명"].get(code, ""), skip_mult
    return None


def get_set_multiplier(option, mapping):
    """
    수량 배수 추출 룰 v2 (2026-04-28 패치).
    --------------------------------------
    감지 패턴 (우선순위 순):
      1. "1+1" / "2+2" / "3+3" → 세트_배수 dict (1+1=1, 2+2=2, 3+3=3)
         - 어차피 옵션에 "1+1"이라고 적혀있으면 1세트 = 2개. ordQty=1 주문 → 출고 2개.
         - "2+2" → 1세트 = 4개. 세트_배수에 2 정의되어 있으면 ordQty=1 → multiplier=2 → 출고 2개? 아니, 4개여야.
         - 따라서 세트_배수 dict 의미는 "1세트당 곱할 단품 수"가 아니라 "set_multiplier" 자체.
      2. "N개 :" / "N개," → N (뉴트리정 할인이벤트)
      3. 일반 "N개" → N
      4. "N+N" 일반형 → N+N 합산
      5. 매칭 없음 → 1
    풀필 수량 = ordQty × get_set_multiplier(option, mapping).
    """
    if not option:
        return 1
    mult_dict = mapping.get("세트_배수", {"1+1": 1, "2+2": 2, "3+3": 3})
    for pat, mult in mult_dict.items():
        if pat in option:
            return mult
    m = re.search(r"(\d+)\s*개\s*[:：,]", option)
    if m:
        return int(m.group(1))
    m = re.search(r"(\d+)\s*개", option)
    if m:
        return int(m.group(1))
    m = re.search(r"(\d+)\+(\d+)", option)
    if m:
        return int(m.group(1)) + int(m.group(2))
    return 1


def check_duplicate_with_history(rows, history_paths=None):
    """
    중복 발주 가드 (2026-04-28 v0.2.1 추가).
    이미 풀필먼트에 등록된 쇼핑몰주문번호가 이번 발주에 또 들어가는지 검사.
    history_paths: 풀필먼트 발주조회 raw JSON 또는 엑셀 변환본 경로들.
    각 history는 [{'쇼핑몰주문번호': ..., '발주등록일': ..., '오더코드': ...}] 형식 기대.

    검사 룰:
      - 같은 shmaOrdNo (주문번호) 가 history 에 이미 있으면 ⚠️ 차단
      - 단, 사용자 명시 강제(force) 가 아니면 업로드 거부
    """
    if not history_paths:
        return []
    history = set()
    for p in history_paths:
        try:
            d = json.load(open(p, "r", encoding="utf-8"))
            if isinstance(d, list):
                for r in d:
                    if isinstance(r, dict):
                        s = str(r.get("쇼핑몰주문번호") or r.get("shmaOrdNo") or r.get("주문번호") or "").strip()
                        if s:
                            history.add(s)
        except Exception:
            continue
    if not history:
        return []
    dups = []
    for r in rows:
        s = str(r.get("주문번호") or "").strip()
        if s and s in history:
            dups.append({
                "shmaOrdNo": s,
                "recv": r.get("받는분 이름"),
                "code": r.get("상품고유코드"),
            })
    return dups


def validate_qty_drops(rows, orders):
    """
    풀필 수량 < ordQty 인데 옵션에 명시적 N개/N+N 표기 없는 경우 = silent drop 의심.
    """
    issues = []
    by_key = {}
    for o in orders:
        k = str(o.get("shmaOrdNo") or o.get("ordNo") or "")
        by_key.setdefault(k, o)
    for r in rows:
        k = str(r.get("주문번호") or "")
        o = by_key.get(k)
        if not o:
            continue
        try:
            ord_qty = int(o.get("qty") or 1)
            ff_qty = int(r.get("수량") or 0)
        except (TypeError, ValueError):
            continue
        if ord_qty > 1 and ff_qty < ord_qty:
            opt = str(o.get("option") or "")
            has_explicit_mult = bool(
                re.search(r"(\d+\+\d+|\d+\s*개|\d+\s*[pP]\b|\d+\s*[Ss]et)", opt)
            )
            if not has_explicit_mult:
                issues.append({
                    "ord": k, "recv": r.get("받는분 이름"),
                    "ordQty": ord_qty, "ffQty": ff_qty,
                    "code": r.get("상품고유코드"), "opt": opt[:60],
                })
    return issues


def parse_sabang_orders(path):
    with open(path, "r", encoding="utf-8") as f:
        raw = json.load(f)
    orders = raw if isinstance(raw, list) else raw.get("orders") or raw.get("data") or []
    out = []
    for o in orders:
        item = {std: o.get(sab, "") for sab, std in SABANG_FIELD_MAP.items()}
        item["product_no"] = ""
        try:
            item["qty"] = int(item.get("qty") or 1)
        except (ValueError, TypeError):
            item["qty"] = 1
        out.append(item)
    return out


def parse_smartstore_xlsx(path):
    try:
        import pandas as pd
    except ImportError:
        print("pandas 설치 필요: pip install pandas", file=sys.stderr)
        sys.exit(1)
    df = pd.read_excel(path, header=1)
    out = []
    for _, row in df.iterrows():
        color, size, set_type = parse_option_string(str(row.get("옵션정보", "")))
        out.append({
            "ordNo": "",
            "shmaOrdNo": str(row.get("상품주문번호", "")),
            "product_no": str(row.get("상품번호", "")),
            "product": str(row.get("상품명", "")),
            "option": str(row.get("옵션정보", "")),
            "option_parsed": {"color": color, "size": size, "set_type": set_type},
            "qty": int(row.get("수량", 1) or 1),
            "receiver": str(row.get("수취인명", "")),
            "phone": str(row.get("수취인연락처1", "")),
            "zipcode": _clean(str(row.get("우편번호", ""))),
            "address": str(row.get("통합배송지", "")),
            "message": _clean(str(row.get("배송메세지", ""))),
        })
    return out


def parse_option_string(opt):
    color, size, set_type = None, None, None
    for p in str(opt).split(" / "):
        p = p.strip()
        if p.startswith("색상:"):
            color = p.split(":", 1)[1].strip()
        elif p.startswith("사이즈:"):
            size = p.split(":", 1)[1].strip()
        elif p.startswith("세트"):
            set_type = p
    return color, size, set_type


def _clean(s):
    return "" if s in ("nan", "None", None) else s


def process_orders(orders, mapping):
    rows, unmapped = [], []
    for i, o in enumerate(orders):
        product = o.get("product", "")
        option = o.get("option", "")
        product_no = o.get("product_no", "")
        parsed = o.get("option_parsed") or {}
        color = parsed.get("color")
        size = parsed.get("size")

        matched = None
        if product_no:
            matched = match_by_product_no(mapping, product_no, color, size)
        if not matched:
            matched = match_by_keyword(mapping, product, option)

        if not matched:
            unmapped.append({
                "idx": i,
                "ordNo": o.get("ordNo", ""),
                "shmaOrdNo": o.get("shmaOrdNo", ""),
                "product": product,
                "option": option,
            })
            rows.append({"상품고유코드": "???", "판매상품명": "", **_ship_row(o, 0)})
            continue

        if len(matched) == 3:
            code, name, skip_mult = matched
        else:
            code, name = matched
            skip_mult = False
        base_qty = int(o.get("qty") or 1)
        qty = base_qty if skip_mult else base_qty * get_set_multiplier(option, mapping)
        rows.append({"상품고유코드": code, "판매상품명": name, **_ship_row(o, qty)})
    return rows, unmapped


def _ship_row(o, qty):
    return {
        "수량": qty,
        "배송방식": "택배",
        "주문자 이름": "",
        "받는분 이름": o.get("receiver", ""),
        "전화번호1": o.get("phone", ""),
        "전화번호2": "",
        "우편번호": o.get("zipcode", ""),
        "주소1": o.get("address", ""),
        "주소2": "",
        "배송메세지": o.get("message", ""),
        "주문번호": o.get("shmaOrdNo", "") or o.get("ordNo", ""),
        "관리메모1": "", "관리메모2": "", "관리메모3": "",
        "관리메모4": "", "관리메모5": "",
        "상품별 메모1": "", "상품별 메모2": "", "상품별 메모3": "",
        "발주 타입": "", "출고희망일": o.get("출고희망일") or _default_ship_date(),
    }


def _default_ship_date():
    """기본 출고희망일: 내일 (YYYY-MM-DD)"""
    from datetime import datetime, timedelta
    return (datetime.now() + timedelta(days=1)).strftime("%Y-%m-%d")


def write_excel(rows, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "발주등록"

    red_font = Font(name="Arial", bold=True, color="FF0000", size=10)
    black_bold = Font(name="Arial", bold=True, size=10)
    data_font = Font(name="Arial", size=10)
    header_fill = PatternFill("solid", fgColor="F2F2F2")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col_idx, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = red_font if h in REQUIRED_HEADERS else black_bold
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for r_idx, row in enumerate(rows, 2):
        for c_idx, h in enumerate(HEADERS, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=row.get(h, ""))
            cell.font = data_font
            cell.border = border

    widths = [16, 45, 6, 8, 12, 12, 16, 16, 10