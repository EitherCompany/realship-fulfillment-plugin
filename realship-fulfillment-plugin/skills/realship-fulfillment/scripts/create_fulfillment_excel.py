#!/usr/bin/env python3
"""v0.4.0 데이터 드리븐 룰 엔진 + fixture 회귀 테스트 + 사용자 confirm 게이트."""
import openpyxl, json, datetime, re, warnings, argparse, sys, os, glob
from collections import Counter, defaultdict
from openpyxl import Workbook

warnings.filterwarnings('ignore')

HEADERS = ['상품고유코드','판매상품명','수량','배송방식','주문자 이름','받는분 이름','전화번호1','전화번호2',
           '우편번호','주소1','주소2','배송메세지','주문번호','관리메모1','관리메모2','관리메모3','관리메모4','관리메모5',
           '상품별 메모1','상품별 메모2','상품별 메모3','발주 타입','출고희망일']

LUMI_CODES = {f'E004003{n:02d}' for n in range(43, 55)}
SR = re.compile(r'(\d{3})\s*-\s*(\d{3})')


# === SKU 매트릭스 빌드 (brand 분리) ===

def find_prev_fulfillment(downloads_dir, today_str):
    """Downloads에서 직전 사이클 풀필먼트 엑셀 자동 탐색 (이더+뉴트리, 오늘 이전 가장 최근).
    파일명 패턴: 풀필먼트_{이더|뉴트리}_YYYYMMDD*.xlsx
    우선순위: 파일명 날짜 (가장 최근) > _FINAL > v숫자 (가장 큼) > mtime"""
    if not downloads_dir or not os.path.isdir(downloads_dir):
        return []
    results = []
    for brand in ['이더', '뉴트리']:
        pattern = os.path.join(downloads_dir, f'풀필먼트_{brand}_*.xlsx')
        files = glob.glob(pattern)
        candidates = []
        for f in files:
            name = os.path.basename(f)
            m = re.search(r'(\d{8})', name)
            date_str = m.group(1) if m else '00000000'
            if date_str >= today_str:
                continue  # 오늘 또는 미래 파일 제외 (자기 자신 매칭 방지)
            v = 99 if '_FINAL' in name else 0
            vm = re.search(r'_v(\d+)', name)
            if vm: v = max(v, int(vm.group(1)))
            mtime = os.path.getmtime(f)
            candidates.append((date_str, v, mtime, f))
        if candidates:
            candidates.sort(reverse=True)
            results.append(candidates[0][3])
    return results


def load_prev_fulfillment_rows(paths):
    """이전 풀필먼트 엑셀들을 row dict 리스트로 로드."""
    rows = []
    for p in paths:
        try:
            wb = openpyxl.load_workbook(p, data_only=True); ws = wb.active
            h = [ws.cell(1,c).value for c in range(1, ws.max_column+1)]
            for r in range(2, ws.max_row+1):
                rows.append({h[c]: ws.cell(r, c+1).value for c in range(len(h)) if h[c]})
        except Exception as e:
            print(f'⚠️ 이전 풀필먼트 엑셀 로드 실패 {p}: {e}')
    return rows


def build_sku_matrix(sku_path):
    wb = openpyxl.load_workbook(sku_path, data_only=True); ws = wb.active
    sku, mat = {}, {}
    for r in range(2, ws.max_row + 1):
        code = ws.cell(r, 3).value
        name = ws.cell(r, 4).value
        if not (code and name and (str(code).startswith('E') or str(code).startswith('N'))):
            continue
        sku[code] = name
        if '구름' in name and '깔창' in name:
            m = SR.search(name)
            if m:
                size = f'{m.group(1)}-{m.group(2)}'
                color = '블랙' if '블랙' in name else '그레이' if '그레이' in name else None
                if color:
                    brand = '루미솔' if '루미솔' in name else '이더'
                    mat[('구름깔창', brand, color, size)] = code
        if ('벌집' in name or '메쉬' in name) and '깔창' in name:
            m = SR.search(name)
            if m:
                mat[('벌집깔창', None, None, f'{m.group(1)}-{m.group(2)}')] = code
        if '양말' in name and '클린인테크' in name:
            for c in ['화이트','그레이','블랙']:
                if f'{c} 20p' in name:
                    mat[('양말', None, c, None)] = code
        if '글램루아' in name:
            sm = re.search(r',\s*(L|M|S|XL)\s*$', name)
            if sm:
                seg = name.split(',')
                if len(seg) >= 3:
                    mat[('글램루아', None, seg[-2].strip(), sm.group(1))] = code
        if '아쿠아슈즈' in name:
            sm = SR.search(name)
            if sm:
                size = f'{sm.group(1)}-{sm.group(2)}'
                # 2026-05-15: 출고상품 엑셀 기준 색 4종만. 네이비/그레이 없음.
                for c in ['블랙','블루','퍼플','피치']:
                    if c in name:
                        mat[('아쿠아슈즈', None, c, size)] = code; break
    return sku, mat


# === 단일 평가 엔진 (충돌 감지) ===
def text_color(text):
    if '블랙' in text or '검은색' in text or '검정' in text: return '블랙'
    if '그레이' in text or '회색' in text: return '그레이'
    return None


def evaluate_rule(rule, text, opt):
    """룰의 trigger 조건만 평가. 매칭되면 True."""
    if 'all' in rule:
        if not all(kw in text for kw in rule['all']): return False
    if 'any' in rule:
        if not any(kw in text for kw in rule['any']): return False
    if 'any_all' in rule:
        if not any(all(kw in text for kw in group) for group in rule['any_all']): return False
    if 'none' in rule:
        if any(kw in text for kw in rule['none']): return False
    return True


def resolve_matrix(rule, text, opt, mat):
    """matrix lookup으로 코드 결정. rule['matrix']['lookup'] 우선, 없으면 SKU master 매트릭스(mat) 폴백."""
    cat = rule['matrix']['category']
    brand = rule['matrix'].get('brand')
    extract = rule['matrix']['extract']
    lookup = rule['matrix'].get('lookup')  # 2026-05-15: 룰 내장 lookup table 우선

    # 글램루아: 색 토큰 set + 사이즈 매칭 (특수 로직)
    if cat == '글램루아':
        sm = re.search(r'사이즈[: ]*([SMLXL]+)', opt or '') or re.search(r'\b(L|M|S|XL)\b', opt or '')
        size = sm.group(1) if sm else None
        opt2 = (opt or '').replace(' ','')
        opt_tok = set(re.findall(r'(블랙|스킨|그레이|화이트)', opt2))
        if size:
            # 1) lookup 우선 — 옵션 색조합 + size 키 매칭
            if lookup:
                for lkey, lcode in lookup.items():
                    parts = lkey.split('|')
                    if len(parts) < 2 or parts[-1] != size: continue
                    color_part = '|'.join(parts[:-1])
                    code_tok = set(re.findall(r'(블랙|스킨|그레이|화이트)', color_part))
                    if opt_tok == code_tok: return lcode
            # 2) mat fallback
            for k, v in mat.items():
                if k[0]!='글램루아' or k[3]!=size: continue
                code_tok = set(re.findall(r'(블랙|스킨|그레이|화이트)', k[2]))
                if opt_tok == code_tok: return v
        return None

    # 색 추출
    color = None
    if 'color' in extract:
        if cat == '구름깔창':
            color = text_color(text)
        elif cat == '아쿠아슈즈':
            color = text_color(text) or _find_in(text, ['블루','퍼플','피치'])
        elif cat == '양말':
            color = _find_in(text, ['화이트','블랙','그레이'])

    # 사이즈 추출
    size = None
    if 'size' in extract:
        m = SR.search(text)
        if m: size = f'{m.group(1)}-{m.group(2)}'

    # 룰 내장 lookup 우선 — 카테고리별 키 형식
    if lookup:
        lkey = None
        if cat in ('구름깔창','아쿠아슈즈') and color and size:
            lkey = f'{color}|{size}'
        elif cat == '벌집깔창' and size:
            lkey = size
        elif cat == '양말' and color:
            lkey = color
        if lkey and lkey in lookup: return lookup[lkey]

    # 폴백: SKU master에서 빌드된 매트릭스
    key = (cat, brand, color, size)
    return mat.get(key)



def _find_in(text, candidates):
    for c in candidates:
        if c in text: return c
    return None


def map_with_rules(prod_name, option, rules, mat):
    """단일 평가 엔진. 우선순위 정렬 후 첫 매칭 룰 적용. 충돌 감지."""
    if not prod_name: return None, None, None
    text = f'{prod_name} {option or ""}'
    
    matches = []
    for rule in sorted(rules, key=lambda r: r.get('priority', 999)):
        if evaluate_rule(rule, text, option):
            if 'code' in rule:
                matches.append((rule, rule['code']))
            elif 'matrix' in rule:
                code = resolve_matrix(rule, text, option, mat)
                if code: matches.append((rule, code))
    
    if not matches: return None, None, None
    # 우선순위 정렬되어 있으니 첫 번째 사용
    rule, code = matches[0]
    # 충돌: 동일 우선순위 다른 룰 매칭 시
    same_priority = [m for m in matches if m[0].get('priority') == rule.get('priority')]
    conflict = same_priority if len(same_priority) > 1 else None
    return code, rule['id'], conflict


# === 수량 룰 ===
def apply_quantity_rules(option, qty_rules):
    if not option: return 1
    s = str(option)
    for r in qty_rules:
        m = re.search(r['regex'], s)
        if m:
            if r.get('match_type') == 'equal_groups':
                if len(m.groups()) >= 2 and m.group(1) == m.group(2):
                    return int(m.group(1))
                continue
            if r['multiplier'] == 'first_group':
                return int(m.group(1))
    return 1


# === 빈박스 / 사이클 ===
def is_pre_cycle(r, cycle_start):
    od = r.get('주문일시(YYYY-MM-DD HH:MM)')
    if isinstance(od, datetime.datetime): return od < cycle_start
    if isinstance(od, str):
        try: return datetime.datetime.strptime(od[:16], '%Y-%m-%d %H:%M') < cycle_start
        except: return False
    return False


def classify_binbox(r, binbox_rules):
    mall = r.get('쇼핑몰명(1)') or ''
    for rule in binbox_rules:
        if rule.get('mall') and rule['mall'] != mall: continue
        field_val = r.get(rule.get('field','')) or ''
        if 'contains' in rule and rule['contains'] in str(field_val): return 'binbox'
        if 'contains_exact' in rule and rule['contains_exact'] in str(field_val): return 'binbox'
    return 'realship'


# === 주소 / fallback ===
def add_dong_ho_comma(addr, pattern):
    if not addr: return addr
    addr = str(addr).strip()
    pat = re.compile(pattern)
    m = pat.search(addr)
    if not m: return addr
    idx = m.start(1)
    if idx == 0 or addr[idx-1] == ',': return addr
    return addr[:idx] + ',' + addr[idx:]


def norm(value, fallback):
    s = str(value or '').strip()
    return s if s and s != 'None' else fallback


def norm_zip(value, fallback):
    """우편번호: 하이픈/공백 제거 + 5자리 zfill (앞자리 0 보존). 풀필먼트는 5자리 신우편번호."""
    s = str(value or '').strip().replace('-', '').replace(' ', '')
    if not s or s == 'None': return fallback
    return s.zfill(5) if len(s) <= 5 else s


# === Fixture 회귀 테스트 ===
def run_fixtures(fixture_path, rules, mat, qty_rules, sku, addr_pattern, binbox_rules):
    if not os.path.exists(fixture_path):
        print(f'⚠️ fixture 파일 없음: {fixture_path}')
        return False
    with open(fixture_path, encoding='utf-8') as f:
        F = json.load(f)
    
    failures = []
    for t in F.get('tests', []):
        code, rid, conf = map_with_rules(t['product'], t['option'], rules, mat)
        qty = 1 * apply_quantity_rules(t['option'], qty_rules)
        if code != t['expected_code'] or qty != t['expected_qty_per_ord']:
            failures.append(f"[{t['id']}] code={code}/{t['expected_code']} qty={qty}/{t['expected_qty_per_ord']}")
    
    for t in F.get('address_comma_tests', []):
        out = add_dong_ho_comma(t['input'], addr_pattern)
        if out != t['expected']:
            failures.append(f"[addr_{t['input'][:30]}] expected={t['expected']} got={out}")
    
    for t in F.get('binbox_classification_tests', []):
        row = {'쇼핑몰명(1)': t['mall'], '수취인주소(4)': t['주소'], '배송메세지': t['msg']}
        result = classify_binbox(row, binbox_rules)
        if result != t['expected']:
            failures.append(f"[binbox_{t['mall']}_{t['msg'][:20]}] expected={t['expected']} got={result}")
    
    if failures:
        print(f'🚨 FIXTURE FAIL ({len(failures)}건):')
        for f_ in failures: print(f'  {f_}')
        return False
    print(f'✅ Fixture 통과 ({len(F.get("tests",[]))} mapping + {len(F.get("address_comma_tests",[]))} address + {len(F.get("binbox_classification_tests",[]))} binbox)')
    return True


# === 풀필먼트 엑셀 생성 ===
def to_xlsx(rows, sku, output_path, addr_pattern, fb, ship_date):
    """합배송 정렬: (받는분, 전화) 그룹화. 같은 그룹은 인접 행으로 정렬 (등장 순서 보존)."""
    def npn(p): return str(p or '').strip().replace('-','').replace(' ','')
    groups, order = {}, []
    for r in rows:
        key = ((r.get('수취인명') or '').strip(), npn(r.get('수취인전화번호1')))
        if key not in groups:
            groups[key] = []; order.append(key)
        groups[key].append(r)
    sorted_rows = []
    for k in order: sorted_rows.extend(groups[k])
    rows = sorted_rows
    
    wb = Workbook(); ws = wb.active
    ws.append(HEADERS)
    for r in rows:
        addr_raw = (r.get('수취인주소(4)') or fb['address']).strip()
        addr = add_dong_ho_comma(addr_raw, addr_pattern) if addr_raw != fb['address'] else addr_raw
        ws.append([
            r['_code'],
            sku.get(r['_code'], r.get('_match_id', '')),
            r.get('_qty', r.get('수량') or 1),
            '택배', '',
            (r.get('수취인명') or '').strip(),
            norm(r.get('수취인전화번호1'), fb['phone']), '',
            norm_zip(r.get('수취인우편번호(1)'), fb['zip']),
            addr, '',
            (r.get('배송메세지') or '').strip(),
            str(r.get('주문번호(쇼핑몰)') or '').strip(),
            '','','','','','','','','', ''  # 출고희망일: 빈칸 (사용자 결정 2026-05-15 — 풀필먼트가 임의 처리)
        ])
        # 우편번호 셀(col 9)을 text format으로 강제 → leading 0 보존
        ws.cell(row=ws.max_row, column=9).number_format = '@'
    wb.save(output_path)


# === 검증 보고서 ===
def build_report(buckets, sku, rules_cfg, dup, conflicts):
    final = buckets['final']
    code_dist = Counter(r['_code'] for r in final)
    total = len(final)
    
    stops = []
    cfg = rules_cfg.get('auto_stop_thresholds', {})
    if len(buckets['unmapped']) > cfg.get('unmapped_count', 0):
        stops.append(f'미매핑 {len(buckets["unmapped"])}건')
    if total > 0 and code_dist:
        top_code, top_count = code_dist.most_common(1)[0]
        pct = top_count / total * 100
        wl = cfg.get('code_whitelist_high_pct', [])
        if pct > cfg.get('single_code_pct_excl_whitelist', 50) and top_code not in wl:
            stops.append(f'코드 {top_code} {pct:.1f}% 몰림')
    lumi = sum(v for k,v in code_dist.items() if k in LUMI_CODES)
    if lumi >= cfg.get('lumisol_count_warning', 1):
        stops.append(f'루미솔 {lumi}건')
    fb = rules_cfg['user_fallback']
    zip_fb = sum(1 for r in final if norm(r.get('수취인우편번호(1)'), fb['zip']) == fb['zip'])
    zip_pct = zip_fb / total * 100 if total else 0
    if zip_pct > cfg.get('zip_fallback_pct', 5):
        stops.append(f'우편번호 fallback {zip_pct:.1f}%')
    if len(dup) > cfg.get('crossvalidate_dup_count', 0):
        stops.append(f'cross-validate 중복 {len(dup)}건')
    if conflicts:
        stops.append(f'룰 충돌 {len(conflicts)}건')
    
    qty_NN, qty_N = defaultdict(int), defaultdict(int)
    for r in final:
        opt = r.get('옵션(수집)') or ''
        m = re.search(r'(\d+)\s*\+\s*(\d+)', opt)
        if m and m.group(1) == m.group(2):
            n = int(m.group(1))
            qty_NN[(f'{n}+{n}', float(r.get('수량') or 1), float(r['_qty']))] += 1
        m = re.search(r'(\d+)개\s*:', opt)
        if m:
            n = int(m.group(1))
            qty_N[(f'{n}개', float(r.get('수량') or 1), float(r['_qty']))] += 1
    
    return {
        'totals': {k: len(v) if isinstance(v, list) else v for k,v in buckets.items() if k != 'rows'},
        'top_codes': code_dist.most_common(10),
        'lumisol_count': lumi,
        'zip_fallback_pct': round(zip_pct, 2),
        'qty_matrix_NN': {f'{k[0]}, ordQty={k[1]}, 풀필={k[2]}': v for k,v in qty_NN.items()},
        'qty_matrix_N': {f'{k[0]}, ordQty={k[1]}, 풀필={k[2]}': v for k,v in qty_N.items()},
        'unmapped_samples': [{'product': r.get('상품명(수집)'), 'option': r.get('옵션(수집)')} for r in buckets['unmapped'][:10]],
        'crossvalidate_dup_count': len(dup),
        'rule_conflicts': conflicts or [],
        'auto_stop_reasons': stops,
        'should_proceed': len(stops) == 0,
    }


# === 메인 ===
def main():
    p = argparse.ArgumentParser()
    p.add_argument('--orders', required=True)
    p.add_argument('--sku-master')
    p.add_argument('--fulfillment-history', nargs='*', help='발주조회 엑셀 (cross-validate, 명시 시 우선)')
    p.add_argument('--prev-fulfillment-dir', help='Downloads 폴더 — 자동으로 직전 사이클 풀필먼트 엑셀 탐색 (--fulfillment-history 미제공 시)')
    p.add_argument('--state', required=True)
    p.add_argument('--cycle-start', required=True)
    p.add_argument('--report-only', action='store_true')
    p.add_argument('--output-ether')
    p.add_argument('--output-nutri')
    p.add_argument('--output-report')
    args = p.parse_args()

    # 룰 + fixture
    here = os.path.dirname(os.path.abspath(__file__))
    with open(os.path.join(here, 'mapping_rules.json'), encoding='utf-8') as f:
        rules_cfg = json.load(f)
    rules = rules_cfg['mapping_rules']
    qty_rules = rules_cfg['quantity_rules']
    binbox_rules = rules_cfg['binbox_rules']
    addr_pattern = rules_cfg['address_format']['comma_pattern']
    fb = rules_cfg['user_fallback']

    sku, mat = build_sku_matrix(args.sku_master) if args.sku_master else ({}, {})

    if mat:
        ok = run_fixtures(os.path.join(here, 'fixture_tests.json'), rules, mat, qty_rules, sku, addr_pattern, binbox_rules)
        if not ok: sys.exit(1)

    wb = openpyxl.load_workbook(args.orders, data_only=True); ws = wb.active
    h = [ws.cell(1, c).value for c in range(1, ws.max_column+1)]
    rows = [{h[c]: ws.cell(r, c+1).value for c in range(len(h)) if h[c]}
            for r in range(2, ws.max_row+1)]

    cycle_start = datetime.datetime.strptime(args.cycle_start, '%Y-%m-%d %H:%M')

    try:
        with open(args.state, encoding='utf-8') as f: state = json.load(f)
    except: state = {}
    prev_pairs = set()
    for k, v in state.items():
        if isinstance(v, dict):
            for plist in [v.get('shma_code_pairs',[]), v.get('ether_code_pairs',[]), v.get('nutri_code_pairs',[])]:
                for p_ in plist:
                    if isinstance(p_, list) and len(p_)==2:
                        prev_pairs.add(tuple(p_))

    pre_cycle, binbox, realship = [], [], []
    for r in rows:
        if is_pre_cycle(r, cycle_start): pre_cycle.append(r); continue
        c = classify_binbox(r, binbox_rules)
        (binbox if c == 'binbox' else realship).append(r)

    mapped, unmapped = [], []
    conflicts = []
    for r in realship:
        code, rid, conf = map_with_rules(r.get('상품명(수집)'), r.get('옵션(수집)'), rules, mat)
        if code:
            r['_code'] = code; r['_match_id'] = rid
            r['_qty'] = (r.get('수량') or 1) * apply_quantity_rules(r.get('옵션(수집)'), qty_rules)
            mapped.append(r)
            if conf: conflicts.append({'row': r.get('수취인명'), 'matched_rules': [c[0]['id'] for c in conf]})
        else:
            unmapped.append(r)

    state_skip, after_state = [], []
    for r in mapped:
        pair = (str(r.get('주문번호(쇼핑몰)','')), r['_code'])
        if pair in prev_pairs: state_skip.append(r)
        else: after_state.append(r)

    fulfill = []
    cv_source = None
    if args.fulfillment_history:
        cv_source = '발주조회 엑셀 (사용자 명시)'
        for fpath in args.fulfillment_history:
            fws = None
            # 1) msoffcrypto 복호화 시도 (암호화된 경우)
            try:
                import msoffcrypto, io
                with open(fpath, 'rb') as f:
                    of = msoffcrypto.OfficeFile(f); of.load_key(password='dlejrhddyd1!')
                    buf = io.BytesIO(); of.decrypt(buf); buf.seek(0)
                fws = openpyxl.load_workbook(buf, data_only=True).active
            except Exception as e:
                # 2) 폴백: 이미 풀려있는 파일이면 직접 열기
                try:
                    fws = openpyxl.load_workbook(fpath, data_only=True).active
                    print(f'ℹ️ 발주조회 폴백 로드 (암호화 X): {os.path.basename(fpath)}')
                except Exception as e2:
                    print(f'⚠️ 발주조회 로드 실패: {e} / 폴백 실패: {e2}')
                    continue
            if fws is None: continue
            fh = [fws.cell(1, c).value for c in range(1, fws.max_column+1)]
            for rr in range(2, fws.max_row+1):
                fulfill.append({fh[c]: fws.cell(rr, c+1).value for c in range(len(fh)) if fh[c]})
    elif args.prev_fulfillment_dir:
        today_str = datetime.date.today().strftime('%Y%m%d')
        prev_files = find_prev_fulfillment(args.prev_fulfillment_dir, today_str)
        if prev_files:
            cv_source = f'어제 풀필먼트 엑셀 자동 탐색 ({len(prev_files)}개)'
            print(f'🔍 자동 cross-validate: {[os.path.basename(p) for p in prev_files]}')
            fulfill = load_prev_fulfillment_rows(prev_files)

    def npn(p): return str(p or '').strip().replace('-','').replace(' ','')
    fA, fB = defaultdict(list), defaultdict(list)
    fOrderCodes = defaultdict(set)  # 2026-05-15: 주문번호 → {이전 사이클의 코드 set} (코드 변동 감지용)
    # 발주조회: 주문번호/고유코드. 어제 풀필먼트 엑셀: 주문번호/상품고유코드.
    for r in fulfill:
        on = str(r.get('주문번호') or '').strip()
        cd = str(r.get('고유코드') or r.get('상품고유코드') or '').strip()
        nm = str(r.get('받는분 이름') or '').strip()
        ph = npn(r.get('전화번호1'))
        if on and cd: fA[(on, cd)].append(r); fOrderCodes[on].add(cd)
        if nm and ph and cd: fB[(nm, ph, cd)].append(r)
    if cv_source: print(f'✅ Cross-validate source: {cv_source} ({len(fulfill)} rows)')

    dup, final, code_mismatch = [], [], []
    for r in after_state:
        on = str(r.get('주문번호(쇼핑몰)','')).strip()
        cd = r['_code']
        kA = (on, cd)
        kB = ((r.get('수취인명') or '').strip(), npn(r.get('수취인전화번호1')), cd)
        if (fA and kA in fA) or (fB and kB in fB):
            dup.append(r)
        else:
            # 코드 변동 감지: 같은 주문번호인데 이전 사이클의 코드 set에 현재 코드가 없음
            if on and on in fOrderCodes and cd not in fOrderCodes[on]:
                code_mismatch.append({
                    '주문번호': on,
                    '현재코드': cd,
                    '이전코드': sorted(fOrderCodes[on]),
                    '상품명': str(r.get('상품명(수집)') or r.get('상품명(확정)') or '')[:50],
                })
            final.append(r)

    ether = [r for r in final if r['_code'].startswith('E')]
    nutri = [r for r in final if r['_code'].startswith('N')]

    buckets = {'rows':rows, 'pre_cycle_skip':pre_cycle, 'binbox':binbox, 'realship':realship,
               'mapped':mapped, 'unmapped':unmapped, 'state_skip':state_skip, 'final':final,
               'ether':ether, 'nutri':nutri}
    report = build_report(buckets, sku, rules_cfg, dup, conflicts)
    report['code_mismatch'] = code_mismatch  # 2026-05-15: 코드 변동 감지 (정보용, stop X)
    report['code_mismatch_count'] = len(code_mismatch)

    if args.output_report:
        with open(args.output_report, 'w', encoding='utf-8') as f:
            json.dump(report, f, ensure_ascii=False, indent=2, default=str)

    print(f'\n[검증 보고서]')
    t = report['totals']
    print(f'  전체 {len(rows)} | 사이클이전 {t["pre_cycle_skip"]} | 빈박스 {t["binbox"]} | 실배송 {t["realship"]}')
    print(f'  매핑 {t["mapped"]} / 미매핑 {t["unmapped"]}')
    print(f'  state SKIP {t["state_skip"]} / cross-validate 중복 {report["crossvalidate_dup_count"]} / 코드 변동 {report.get("code_mismatch_count",0)}')
    if code_mismatch:
        print(f'  ⚠️ 코드 변동 감지 (이전 사이클과 다른 코드로 매핑됨):')
        for m in code_mismatch[:5]: print(f'    주문 {m["주문번호"]}: {m["이전코드"]} → {m["현재코드"]} ({m["상품명"]})')
    print(f'  최종 {t["final"]} (이더 {t["ether"]} / 뉴트리 {t["nutri"]})')
    print(f'  루미솔 {report["lumisol_count"]} | 우편 fallback {report["zip_fallback_pct"]}%')
    print(f'  코드 TOP 5: {report["top_codes"][:5]}')
    if conflicts:
        print(f'  ⚠️ 룰 충돌 {len(conflicts)}건')
    print(f'\n자동 stop: {len(report["auto_stop_reasons"])}건')
    for s in report["auto_stop_reasons"]: print(f'  ⚠️ {s}')

    if args.report_only:
        print('\n📋 --report-only — 풀필먼트 엑셀 생성 안 함. 사용자 confirm 후 다시 실행.')
        return

    if not report['should_proceed']:
        print('\n🚨 자동 stop 조건. ABORT.')
        sys.exit(2)

    if args.output_ether: to_xlsx(ether, sku, args.output_ether, addr_pattern, fb, '')
    if args.output_nutri: to_xlsx(nutri, sku, args.output_nutri, addr_pattern, fb, '')

    today_key = datetime.date.today().isoformat()
    state['last_cycle_date'] = today_key
    state[today_key] = {
        'ether_code_pairs': [[str(r.get('주문번호(쇼핑몰)','')), r['_code']] for r in ether],
        'nutri_code_pairs': [[str(r.get('주문번호(쇼핑몰)','')), r['_code']] for r in nutri],
        'shma_code_pairs': [[str(r.get('주문번호(쇼핑몰)','')), r['_code']] for r in final],
        'binbox_skip_count': len(binbox),
        'pre_cycle_skip_count': len(pre_cycle),
        'unmapped_count': len(unmapped),
        'crossvalidate_dup_count': len(dup),
        'rule_conflicts': len(conflicts),
        'cycle_start': args.cycle_start,
        'complete_time': datetime.datetime.now().strftime('%Y-%m-%d %H:%M KST'),
    }
    with open(args.state, 'w', encoding='utf-8') as f:
        json.dump(state, f, ensure_ascii=False, indent=2)


if __name__ == '__main__':
    main()
